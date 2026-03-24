#!/usr/bin/env python3
"""
Vermont Baseball League - Fantasy Baseball Data Processor

Run this script to generate/refresh js/data.js from the Fantrax CSV files.

Usage (run from the repo root directory):
    python3 process_data.py

When adding a new season:
  1. Download the Fantrax standings CSV for that season
  2. Place it in the data/ folder (same folder as the other CSVs)
  3. Add an entry to YEAR_FILES below
  4. Run this script
  5. Commit and push to GitHub — the site updates automatically
"""

import csv
import json
import re
from datetime import datetime
from pathlib import Path

# CSV files live in the data/ subfolder
CSV_DIR = Path(__file__).parent / "data"
OUTPUT_FILE = Path(__file__).parent / "js" / "data.js"

# ── Year → CSV filename mapping ──────────────────────────────────────────────
YEAR_FILES = {
    2011: "2011-Fantrax-Standings-Montana State.csv",
    2012: "2012-Fantrax-Standings-Montana State.csv",
    2013: "2013-Fantrax-Standings-Vermont Baseball.csv",
    2014: "2014-Fantrax-Standings-Vermont Baseball.csv",
    2015: "2015-Fantrax-Standings-Vermont Baseball.csv",
    2016: "2016-Fantrax-Standings-Vermont Baseball.csv",
    2017: "2017-Fantrax-Standings-Vermont Baseball.csv",
    2018: "2018-Fantrax-Standings-Vermont Baseball.csv",
    2019: "2019-Fantrax-Standings-Vermont Baseball.csv",
    2020: "2020-Fantrax-Standings-Vermont Baseball.csv",
    2021: "2021-Fantrax-Standings-Vermont Baseball.csv",
    2022: "2022-Fantrax-Standings-Vermont Baseball.csv",
    2023: "2023-Fantrax-Standings-Vermont Baseball.csv",
    2024: "2024-Fantrax-Standings-Vermont Baseball.csv",
    2025: "2025-Fantrax-Standings-Vermont Baseball.csv",
    # 2026: "2026-Fantrax-Standings-Vermont Baseball.csv",  ← add next year
}

OWNERS_FILE = "League Owners By Year - Sheet1.csv"

# Historical Excel file (pre-Fantrax seasons 2001-2010)
EXCEL_FILE = "fantasybaseballstats_1999-2021.xlsx"

YEAR_LEAGUE_NAME = {2011: "Montana State", 2012: "Montana State"}
DEFAULT_LEAGUE_NAME = "Vermont Baseball"
# NOTE: Update this if you know the league name before 2011
PRE_FANTRAX_LEAGUE_NAME = "Montana State"

# Fallback: username → display name for owners who marked profile "Private"
USERNAME_TO_NAME = {
    "rcramer":        "Robert Cramer Jr",
    "rcramerjr":      "Robert Cramer Jr",
    "raberg1":        "Robert Berg",
    "eberg76":        "Erik Berg",
    "eberg":          "Erik Berg",
    "royals_fan8307": "Adam Miller",
    "robbiemoy13":    "Robbie Moy",
    "noahwolfe":      "Noah Wolfe",
    "dthiebes80":     "David Thiebes",
    "johnvlangos":    "John Vlangos",
    "Gnashers7":      "Bob Nash",
    "spiritualized50":"Eric Tarcha",
    "amazinglarry":   "Jim Daley",
    "jpilon":         "John Pilon",
    "csmartin43":     "Christopher Martinez",
    "Willger":        "Sven Willger",
    "tyadams7":       "Ty Adams",
    "desilvajerry":   "Jerry DeSilva",
    "ejohnson523":    "Erik Johnson",
    "csuram89":       "csuram89",
    "elaguamala":     "Daniel Caveney",
}

# Normalize owner names from the historical Excel spreadsheet
HISTORICAL_OWNER_NAMES = {
    "Robb Cramer":          "Robert Cramer Jr",
    "Bobb Cramer":          "Robert Cramer Jr",
    "Robert Cramer":        "Robert Cramer Jr",
    "Rob Berg":             "Robert Berg",
    "Rob berg":             "Robert Berg",
    "Dan Caveney":          "Daniel Caveney",
    "Chris Martinez":       "Christopher Martinez",
    "Chris martinez":       "Christopher Martinez",
    "Dean Mitchel":         "Dean Mitchell",   # typo in 2008
    "Erik berg":            "Erik Berg",
    "Jim Daley ?":          "Jim Daley",
}

# Columns parsed as numbers
NUMERIC_COLS = {
    "Rk", "FPts", "FP/G", "GP", "Hit", "Pit", "WW", "PBL",
    "H", "R", "2B", "3B", "HR", "RBI", "BB", "SO", "SB", "CS",
    "HBP", "SH", "SF", "GIDP", "CYC", "E", "AOF",
    "IP", "W", "L", "SV", "BS", "CG", "ER", "K", "BK", "WP",
    "HB", "SHO", "NH", "PG",
}

# Values in the Fantrax owners export that are labels / meta-data, not team names
FANTRAX_LABELS = {
    "Username", "Real Name", "Email", "Phone", "Last Used", "Days",
    "Select all", "Send email", "Subject", "Message",
    "Please select team owners above.", "Team Managers",
    "Private", "",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_num(s):
    """Parse a possibly comma-formatted number string to float."""
    if not s or not s.strip():
        return None
    try:
        return float(s.strip().replace(",", ""))
    except ValueError:
        return None


def read_csv(filepath):
    with open(filepath, "r", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def parse_standings_sections(filepath):
    """
    Split a Fantrax standings CSV into named sections.
    Returns:  {section_name: {header: [...], rows: [[...]]}}
    """
    rows = read_csv(filepath)
    sections = {}
    cur_section = None
    cur_header = None
    cur_rows = []

    for row in rows:
        # Blank row → end of section
        if not row or all(c.strip() == "" for c in row):
            if cur_section and cur_header and cur_rows:
                sections[cur_section] = {"header": cur_header, "rows": cur_rows}
            cur_section = None
            cur_header = None
            cur_rows = []
            continue

        if cur_section is None:
            cur_section = row[0].strip()
            continue
        if cur_header is None:
            cur_header = [c.strip() for c in row]
            continue
        cur_rows.append(row)

    # Final section (no trailing blank line)
    if cur_section and cur_header and cur_rows:
        sections[cur_section] = {"header": cur_header, "rows": cur_rows}

    return sections


def section_to_list(section):
    header = section["header"]
    result = []
    for row in section["rows"]:
        entry = {}
        for i, col in enumerate(header):
            raw = row[i].strip() if i < len(row) else ""
            entry[col] = parse_num(raw) if col in NUMERIC_COLS else raw
        result.append(entry)
    return result


def process_year(year, filename):
    filepath = CSV_DIR / filename
    if not filepath.exists():
        print(f"  WARNING: {filepath.name} not found — skipping {year}.")
        return None

    sections = parse_standings_sections(filepath)

    data = {
        "year": year,
        "league": YEAR_LEAGUE_NAME.get(year, DEFAULT_LEAGUE_NAME),
    }

    for key, name in [
        ("standings",    "Standings"),
        ("hittingPts",   "Standings - Points - Hitting"),
        ("pitchingPts",  "Standings - Points - Pitching"),
        ("hittingStats", "Standings - Statistics - Hitting"),
        ("pitchingStats","Standings - Statistics - Pitching"),
    ]:
        data[key] = section_to_list(sections[name]) if name in sections else []

    return data


# ── Owners CSV parser ─────────────────────────────────────────────────────────

def is_skip_value(val):
    """True if this cell is a label, date, plain number, email, or phone."""
    v = val.strip()
    if not v or v in FANTRAX_LABELS:
        return True
    # Date: "Oct 18, 2014"
    if re.match(r"^[A-Za-z]+\.?\s+\d{1,2},\s*\d{4}$", v):
        return True
    # Small integer (days count): 1-3 digits
    if re.match(r"^\d{1,3}$", v):
        return True
    # Phone number (7-11 digits)
    if re.match(r"^\d{7,11}$", v):
        return True
    # Comma-formatted number: "1,234"
    if re.match(r"^\d{1,3},\d{3}$", v):
        return True
    # Email
    if "@" in v:
        return True
    return False


def parse_owners(valid_teams_by_year=None):
    """
    Parse the Fantrax owners CSV.
    valid_teams_by_year: {year: set(team_name)} — used to filter out garbage.
    Returns: {owner_display_name: {year: team_name}}
    """
    filepath = CSV_DIR / OWNERS_FILE
    if not filepath.exists():
        print(f"  WARNING: Owners file not found — owner data will be empty.")
        return {}

    rows = read_csv(filepath)
    years = list(range(2011, 2026))
    valid = valid_teams_by_year or {}

    owner_map = {}   # {display_name: {year: team_name}}

    # Real data starts at row index 10 (skip header block)
    data_rows = rows[10:]

    for col_idx, year in enumerate(years):
        # Skip years with no standings data
        if valid and year not in valid:
            continue
        csv_col = col_idx + 1          # column 0 is the label column
        col_vals = [
            (r[csv_col].strip() if csv_col < len(r) else "")
            for r in data_rows
        ]

        seen_teams = set()
        i = 0
        while i < len(col_vals):
            val = col_vals[i]

            if not is_skip_value(val):
                # Candidate team name — look ahead for Username / Real Name
                team_name = val
                username = None
                real_name = None
                after_user_lbl = False
                after_name_lbl = False

                for j in range(i + 1, min(i + 13, len(col_vals))):
                    jv = col_vals[j].strip()

                    if after_user_lbl:
                        if jv and not is_skip_value(jv):
                            username = jv
                        after_user_lbl = False

                    elif after_name_lbl:
                        if jv and jv not in {"Private", ""}:
                            real_name = jv
                        after_name_lbl = False
                        break   # Got what we need

                    elif jv == "Username":
                        after_user_lbl = True
                    elif jv == "Real Name":
                        after_name_lbl = True

                # Resolve display name
                display = None
                if real_name and not is_skip_value(real_name):
                    display = real_name
                elif username:
                    display = USERNAME_TO_NAME.get(username, username)

                # Validate team_name against known standings teams for this year
                known = valid.get(year, set())
                is_valid_team = (not known) or (team_name in known)

                if display and team_name and team_name not in seen_teams and is_valid_team:
                    seen_teams.add(team_name)
                    if display not in owner_map:
                        owner_map[display] = {}
                    if year not in owner_map[display]:
                        owner_map[display][year] = team_name

            i += 1

    return owner_map


# ── Historical Excel parser (pre-Fantrax seasons 2001–2010) ──────────────────

def normalize_historical_owner(name):
    if not name:
        return None
    name = name.strip()
    return HISTORICAL_OWNER_NAMES.get(name, name)


def parse_excel_history():
    """
    Parse pre-Fantrax season data from the historical Excel spreadsheet.
    Only processes years not already covered by YEAR_FILES.
    Returns: (seasons_dict, hist_owners_dict)
    """
    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed — skipping historical Excel data.")
        print("           Run: pip3 install openpyxl")
        return {}, {}

    filepath = CSV_DIR / EXCEL_FILE
    if not filepath.exists():
        print(f"  WARNING: {EXCEL_FILE} not found — skipping historical data.")
        return {}, {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    seasons = {}
    hist_owners = {}   # {owner_display_name: {year: team_name}}

    existing_years = set(YEAR_FILES.keys())

    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        # Skip years already covered by Fantrax CSVs, and pre-2001 (no stats)
        if year in existing_years or year < 2001:
            continue

        ws = wb[sheet_name]
        all_rows = [r for r in ws.iter_rows(values_only=True)
                    if any(v is not None for v in r)]
        if not all_rows:
            continue

        standings = []

        if year == 2009:
            # Double header: row 0 = section labels, row 1 = column names, row 2+ = data
            # Columns: Rank(0), TEAM(1), Total FP(2), G(3), AB(4),
            #          Hitters FP(5), FP/G(6), G(7), SPG(8), RPG(9), IP(10),
            #          Pitchers FP(11), FP/G(12)
            data_rows = all_rows[2:]
            for i, row in enumerate(data_rows):
                team_raw = row[1] if len(row) > 1 else None
                if not team_raw:
                    break
                team = str(team_raw).strip()
                if not team or "league note" in team.lower():
                    break
                rk   = row[0] if isinstance(row[0], (int, float)) else (i + 1)
                # Rank resetting to 1 means a new table has started — stop
                if rk == 1 and standings:
                    break
                fp   = row[2]
                hit  = row[5]
                pit  = row[11] if len(row) > 11 else None
                if not isinstance(fp, (int, float)):
                    continue
                standings.append({
                    "Rk":   float(rk),
                    "Team": team,
                    "FPts": float(fp),
                    "Hit":  float(hit) if isinstance(hit, (int, float)) else None,
                    "Pit":  float(pit) if isinstance(pit, (int, float)) else None,
                })
            # No owner column in 2009 sheet

        else:
            # Standard format: row 0 = header, row 1+ = data
            header = [str(v).lower().strip() if v else "" for v in all_rows[0]]

            def find_col(*names):
                for name in names:
                    for i, h in enumerate(header):
                        if h == name.lower():
                            return i
                return None

            team_col  = find_col("team")
            owner_col = find_col("owner")
            pts_col   = find_col("total points", "total fp")

            if team_col is None or pts_col is None:
                continue

            for i, row in enumerate(all_rows[1:]):
                team_raw = row[team_col] if team_col < len(row) else None
                if not team_raw:
                    break
                team = str(team_raw).strip()
                if not team or "league note" in team.lower():
                    break

                fp_raw = row[pts_col] if pts_col < len(row) else None
                if not isinstance(fp_raw, (int, float)):
                    continue

                owner_raw = row[owner_col] if (owner_col is not None and owner_col < len(row)) else None
                owner = normalize_historical_owner(str(owner_raw).strip()) if owner_raw else None

                standings.append({
                    "Rk":   float(i + 1),
                    "Team": team,
                    "FPts": float(fp_raw),
                    "Hit":  None,
                    "Pit":  None,
                })

                if owner and team:
                    if owner not in hist_owners:
                        hist_owners[owner] = {}
                    if year not in hist_owners[owner]:
                        hist_owners[owner][year] = team

        if standings:
            league = YEAR_LEAGUE_NAME.get(year, PRE_FANTRAX_LEAGUE_NAME)
            seasons[year] = {
                "year":          year,
                "league":        league,
                "standings":     standings,
                "hittingPts":    [],
                "pitchingPts":   [],
                "hittingStats":  [],
                "pitchingStats": [],
            }
            print(f"  {year}  (historical — {len(standings)} teams)")

    return seasons, hist_owners


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Vermont Baseball League — processing CSV data")
    print("=" * 50)

    # Fantrax seasons (2011+)
    all_seasons = {}
    for year, filename in sorted(YEAR_FILES.items()):
        print(f"  {year}  {filename}")
        data = process_year(year, filename)
        if data:
            all_seasons[year] = data

    # Historical Excel seasons (2001–2010)
    print()
    print("  Parsing historical Excel data...")
    excel_seasons, excel_owners = parse_excel_history()
    for year, data in excel_seasons.items():
        if year not in all_seasons:
            all_seasons[year] = data

    # Build per-year set of valid team names for owner-data validation
    valid_teams_by_year = {
        year: {r["Team"] for r in data["standings"] if r.get("Team")}
        for year, data in all_seasons.items()
    }

    print()
    print("  Parsing owners...")
    owners = parse_owners(valid_teams_by_year)

    # Merge historical owner data (Excel) into owners dict
    for name, by_year in excel_owners.items():
        if name not in owners:
            owners[name] = {}
        for year, team in by_year.items():
            if year not in owners[name]:
                owners[name][year] = team

    now = datetime.now().isoformat(timespec="seconds")
    output = (
        f"// Auto-generated by process_data.py — do not edit directly.\n"
        f"// Re-run process_data.py to refresh from CSV files.\n"
        f"// Generated: {now}\n\n"
        f"const LEAGUE_DATA = {json.dumps(all_seasons, indent=2)};\n\n"
        f"const OWNER_DATA = {json.dumps(owners, indent=2)};\n"
    )

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(output, encoding="utf-8")

    print()
    print(f"  ✓  {OUTPUT_FILE}")
    print(f"  ✓  {len(all_seasons)} seasons  |  {len(owners)} owners")
    print()
    print("Open index.html in your browser to view the app.")


if __name__ == "__main__":
    main()
